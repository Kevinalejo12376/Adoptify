# -*- coding: utf-8 -*-
"""Prueba de integracion del reporte de solicitudes del USUARIO (PDF/Excel).

Valida:
  - Usuario con solicitudes -> PDF y Excel con status 200 (datos reales).
  - El Excel contiene SOLO las solicitudes del usuario (aislamiento).
  - Usuario sin solicitudes -> 404 (descarga vacia bloqueada).
  - Sin token -> 401.
"""
from fastapi.testclient import TestClient

from app.main import app
from app.core.security import create_access_token
from app.db.database import SessionLocal
from app.models.usuario import Usuario
from app.models.solicitud import SolicitudAdopcion
from app.models.catalogos import Rol

client = TestClient(app)


def _token(email):
    return create_access_token({"sub": email})


db = SessionLocal()

# --- 1) Usuario con solicitudes -> 200 PDF/Excel ---
usuarios_con = (
    db.query(Usuario)
    .join(SolicitudAdopcion, SolicitudAdopcion.usuario_id == Usuario.id)
    .all()
)
seen, u_con = set(), []
for u in usuarios_con:
    if u.id not in seen:
        seen.add(u.id)
        u_con.append(u)
print("Usuarios con solicitudes:", [(u.id, u.email) for u in u_con[:5]])

if u_con:
    u = u_con[0]
    ids = [s.id for s in db.query(SolicitudAdopcion)
           .filter(SolicitudAdopcion.usuario_id == u.id).all()]
    print(f"\n=== Usuario {u.email} — solicitudes reales={sorted(ids)} ===")
    for fmt, ext in [("pdf", ".pdf"), ("excel", ".xlsx")]:
        resp = client.get(
            f"/api/solicitudes/mias/reporte?formato={fmt}",
            headers={"Authorization": f"Bearer {_token(u.email)}"},
        )
        print(f"  formato={fmt} status={resp.status_code} "
              f"ctype={resp.headers.get('content-type')} "
              f"cd={resp.headers.get('content-disposition')}")
        if resp.status_code == 200:
            nombre = f"prueba_reporte_usuario{ext}"
            open(nombre, "wb").write(resp.content)
            print(f"    -> {nombre} ({len(resp.content)} bytes)")
            if ext == ".xlsx":
                from openpyxl import load_workbook
                wb = load_workbook(nombre)
                ws = wb.active
                ids_excel = [int(r[0].value) for r in ws.iter_rows(min_row=5) if r[0].value]
                ok = set(ids_excel) == set(ids)
                print(f"    -> IDs en el Excel: {sorted(ids_excel)} | "
                      f"contiene SOLO las del usuario: {ok}")
else:
    print("\n(no hay usuario con solicitudes en la BD local)")

# --- 2) Usuario sin solicitudes -> 404 ---
u_sin = (
    db.query(Usuario).join(Rol, Rol.id == Usuario.rol_id)
    .filter(Rol.codigo == "usuario")
    .outerjoin(SolicitudAdopcion, SolicitudAdopcion.usuario_id == Usuario.id)
    .filter(SolicitudAdopcion.id.is_(None))
    .first()
)
print("\n=== Usuario SIN solicitudes -> 404 ===")
if u_sin:
    resp = client.get(
        "/api/solicitudes/mias/reporte?formato=pdf",
        headers={"Authorization": f"Bearer {_token(u_sin.email)}"},
    )
    print(f"  user={u_sin.email} status={resp.status_code} "
          f"detail={resp.json().get('detail')}")
else:
    print("  (no hay usuario sin solicitudes en la BD local)")

# --- 3) Sin token -> 401 ---
print("\n=== Sin token -> 401 ===")
resp = client.get("/api/solicitudes/mias/reporte?formato=pdf")
print(f"  status={resp.status_code}")

db.close()
print("\n[FIN]")
