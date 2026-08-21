# -*- coding: utf-8 -*-
"""Prueba de integracion del reporte de solicitudes por refugio (PDF/Excel).

Valida:
  A) Refugio con solicitudes -> PDF y Excel con status 200.
  B) Refugio sin solicitudes -> 404 (no permite descarga vacia).
  C) Usuario normal (no refugio) -> 403.
  D) Sin token -> 401.
  E) Aislamiento: el Excel del refugio solo contiene SUS solicitudes.
"""
from fastapi.testclient import TestClient

from app.main import app
from app.core.security import create_access_token
from app.db.database import SessionLocal
from app.models.usuario import Usuario
from app.models.refugio import Refugio
from app.models.mascota import Mascota
from app.models.solicitud import SolicitudAdopcion
from app.models.catalogos import Rol

client = TestClient(app)


def _token(email):
    return create_access_token({"sub": email})


db = SessionLocal()

# --- Identificar refugios con / sin solicitudes ---
refugios_info = []
for u in db.query(Usuario).join(Refugio, Refugio.usuario_id == Usuario.id).all():
    r = db.query(Refugio).filter(Refugio.usuario_id == u.id).first()
    mascotas = db.query(Mascota).filter(Mascota.refugio_id == r.id).all()
    ids = [m.id for m in mascotas]
    n_sols = (
        db.query(SolicitudAdopcion)
        .filter(SolicitudAdopcion.mascota_id.in_(ids)).count()
        if ids else 0
    )
    refugios_info.append((u, r, ids, n_sols))
    print(f"Refugio user={u.id} ({u.email}) refugio={r.id} "
          f"mascotas={len(ids)} solicitudes={n_sols}")

con_sols = [x for x in refugios_info if x[3] > 0]
sin_sols = [x for x in refugios_info if x[3] == 0]
usuario_normal = (
    db.query(Usuario).join(Rol, Rol.id == Usuario.rol_id)
    .filter(Rol.codigo == "usuario").first()
)

print("\n=== A) Refugio CON solicitudes -> PDF y Excel ===")
if con_sols:
    u, r, mascota_ids, _ = con_sols[0]
    # IDs reales de solicitudes de ESTE refugio
    reales = {
        s.id for s in db.query(SolicitudAdopcion)
        .filter(SolicitudAdopcion.mascota_id.in_(mascota_ids)).all()
    }
    print(f"  Refugio {r.id}: solicitudes reales={sorted(reales)}")
    for fmt, ext in [("pdf", ".pdf"), ("excel", ".xlsx")]:
        resp = client.get(
            f"/api/solicitudes/recibidas/reporte?formato={fmt}",
            headers={"Authorization": f"Bearer {_token(u.email)}"},
        )
        print(f"  formato={fmt} status={resp.status_code} "
              f"ctype={resp.headers.get('content-type')} "
              f"cd={resp.headers.get('content-disposition')}")
        if resp.status_code == 200:
            nombre = f"prueba_reporte_refugio{ext}"
            open(nombre, "wb").write(resp.content)
            print(f"    -> {nombre} ({len(resp.content)} bytes)")
            if ext == ".xlsx":
                # Aislamiento: abrir el Excel y comparar IDs con los reales
                from openpyxl import load_workbook
                wb = load_workbook(nombre)
                ws = wb.active
                ids_excel = [int(row[0].value) for row in ws.iter_rows(min_row=5) if row[0].value]
                print(f"    -> IDs en el Excel: {sorted(ids_excel)}")
                if set(ids_excel) == reales:
                    print("    -> OK: el Excel contiene EXACTAMENTE las solicitudes del refugio")
                else:
                    print("    -> ERROR: el Excel NO coincide con las solicitudes del refugio")
else:
    print("  (no hay refugio con solicitudes en la BD local)")

print("\n=== B) Refugio SIN solicitudes -> 404 (descarga vacia bloqueada) ===")
if sin_sols:
    u, r, _, _ = sin_sols[0]
    resp = client.get(
        "/api/solicitudes/recibidas/reporte?formato=pdf",
        headers={"Authorization": f"Bearer {_token(u.email)}"},
    )
    print(f"  status={resp.status_code} detail={resp.json().get('detail')}")
else:
    print("  (no hay refugio sin solicitudes; se valida por codigo 404 del endpoint)")

print("\n=== C) Usuario normal (no refugio) -> 403 ===")
if usuario_normal:
    resp = client.get(
        "/api/solicitudes/recibidas/reporte?formato=pdf",
        headers={"Authorization": f"Bearer {_token(usuario_normal.email)}"},
    )
    print(f"  user={usuario_normal.email} status={resp.status_code} "
          f"detail={resp.json().get('detail')}")
else:
    print("  (no hay usuario normal en la BD local)")

print("\n=== D) Sin token -> 401 ===")
resp = client.get("/api/solicitudes/recibidas/reporte?formato=pdf")
print(f"  status={resp.status_code}")

print("\n=== E) Aislamiento entre refugios ===")
if len(con_sols) >= 2:
    u1, r1, ids1, _ = con_sols[0]
    u2, r2, ids2, _ = con_sols[1]
    reales2 = {
        s.id for s in db.query(SolicitudAdopcion)
        .filter(SolicitudAdopcion.mascota_id.in_(ids2)).all()
    }
    resp = client.get(
        "/api/solicitudes/recibidas/reporte?formato=excel",
        headers={"Authorization": f"Bearer {_token(u1.email)}"},
    )
    open("prueba_reporte_refugio_aislamiento.xlsx", "wb").write(resp.content)
    from openpyxl import load_workbook
    wb = load_workbook("prueba_reporte_refugio_aislamiento.xlsx")
    ws = wb.active
    ids_excel1 = [int(row[0].value) for row in ws.iter_rows(min_row=5) if row[0].value]
    filtro = set(ids_excel1).isdisjoint(reales2)
    print(f"  Refugio1={r1.id} IDs_excel={sorted(ids_excel1)} | Refugio2={r2.id} IDs_reales={sorted(reales2)}")
    print(f"  -> No comparte solicitudes con el otro refugio: {filtro}")
else:
    print("  (se requieren 2 refugios con solicitudes para esta validacion)")

db.close()
print("\n[FIN]")
