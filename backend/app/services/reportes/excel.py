"""
Generacion de reportes Excel con openpyxl.

El libro se construye completamente en memoria (``io.BytesIO``) y se devuelve
como bytes. Se aplica formato profesional y sutil: logo institucional bien
acomodado, bloque de metadatos (usuario que genera, fecha y total de registros),
encabezado con relleno suave de la marca, autofiltro, fila congelada, anchos de
columna, formatos numericos/de fecha, bordes suaves, alturas de fila calculadas
para no cortar el texto y alternancia de filas en tonos muy claros.
"""
# pyrefly: ignore [missing-import]
import math
from datetime import datetime, timezone
from io import BytesIO
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.services.reportes.base import (
    Columna,
    TIPO_ENTERO,
    TIPO_FECHA,
    TIPO_FECHA_HORA,
    TIPO_MONEDA,
    TIPO_NUMERO,
    extraer_datos_usuario,
    HEX_BORDE,
    HEX_ENCABEZADO_TEXTO,
    HEX_PRIMARIO,
    HEX_ROSA_100,
    HEX_ROSA_50,
    HEX_TEXTO,
    HEX_TEXTO_SUAVE,
)

# ---------------------------------------------------------------------------
# Paleta sutil (identidad Adoptify, misma del frontend pero suavizada)
# ---------------------------------------------------------------------------
COLOR_PRIMARIO = HEX_PRIMARIO  # Rose-500 (acentos moderados)
COLOR_ROSA_800 = HEX_ENCABEZADO_TEXTO  # Rose-800 (titulos y encabezados)
COLOR_ROSA_100 = HEX_ROSA_100  # Rose-100 (encabezado de tabla)
COLOR_ROSA_50 = HEX_ROSA_50  # Rose-50 (zebra y banda de metadatos)
COLOR_ROSA_300 = "FDA4AF"  # Rose-300 (linea divisoria de marca)
COLOR_TEXTO = HEX_TEXTO  # Gray-800
COLOR_TEXTO_SUAVE = HEX_TEXTO_SUAVE  # Gray-500
COLOR_BORDE = HEX_BORDE  # Borde rosado muy suave
COLOR_BLANCO = "FFFFFF"

# Columna del logo (A). La tabla de datos comienza en la columna B para que el
# logo tenga su propio espacio sin solaparse con el titulo.
COLUMNA_LOGO = 1
OFFSET = COLUMNA_LOGO + 1

# Filas del encabezado del documento
FILA_TITULO = 1
FILA_SUBTITULO = 2
FILA_METADATOS = 3
# Fila del encabezado de columnas (tras una fila de separacion en la 5)
FILA_ENCABEZADO = 6

_ALINEACIONES = {
    "left": "left",
    "center": "center",
    "right": "right",
}

# Formato openpyxl por defecto segun el tipo de columna
_FORMATOS_EXCEL = {
    TIPO_ENTERO: "#,##0",
    TIPO_NUMERO: "#,##0.00",
    # Moneda (COP): sin centavos -> entero. El punto de miles es solo formato.
    TIPO_MONEDA: '"$"#,##0',
    TIPO_FECHA: "DD/MM/YYYY",
    TIPO_FECHA_HORA: "DD/MM/YYYY HH:MM",
}


def _formato_para(col: Columna) -> str:
    """Determina el formato de numero/fecha de la columna en Excel."""
    if col.formato_excel:
        return col.formato_excel
    return _FORMATOS_EXCEL.get(col.tipo, "General")


def _agregar_logo(hoja, fila: int = FILA_TITULO, columna: int = COLUMNA_LOGO) -> None:
    """Inserta el logo de Adoptify (desde Cloudinary) en la celda indicada.

    Prepara tambien la columna y la fila del logo (ancho/alto) para que quede
    bien acomodado y no se solape con el titulo. Si la descarga o la carga de la
    imagen falla, se omite sin romper el libro Excel.
    """
    from openpyxl.drawing.image import Image as XLImage

    from app.services.reportes.base import obtener_logo_bytes

    logo_bytes = obtener_logo_bytes()
    if logo_bytes is None:
        return
    try:
        from PIL import Image as PILImage

        with PILImage.open(BytesIO(logo_bytes)) as im:
            ancho, alto = im.size
        alto_px = 64
        ancho_px = int(alto_px * ancho / alto)
        img = XLImage(BytesIO(logo_bytes))
        img.width = ancho_px
        img.height = alto_px
        hoja.add_image(img, f"{get_column_letter(columna)}{fila}")
        # Dimensiones de la columna/fila del logo para una distribucion limpia
        hoja.column_dimensions[get_column_letter(columna)].width = 22
        hoja.row_dimensions[fila].height = max(
            hoja.row_dimensions[fila].height or 0, alto_px + 12
        )
    except Exception:
        # Si el logo no se puede cargar, se omite sin romper el reporte.
        pass


def _altura_fila(anchos: List[float], fila_datos: Dict[str, Any], columnas: List[Columna]) -> float:
    """Estima la altura (en puntos) que necesita una fila para no cortar texto.

    Se aproximan las lineas que ocupara cada celda con ajuste de texto a partir
    del ancho de su columna y de la longitud de su contenido, y se devuelve una
    altura proporcional (con un tope para filas extremadamente largas).
    """
    lineas = 1
    for j, col in enumerate(columnas, start=1):
        ancho = anchos[j - 1]
        valor = _valor_celda(fila_datos, col)
        if valor is None:
            continue
        texto = str(valor)
        if not texto:
            continue
        # Un caracter de Calibri 10 ocupa ~1.05 unidades del ancho de columna
        chars_por_linea = max(int(ancho * 1.05), 6)
        lineas_texto = math.ceil(len(texto) / chars_por_linea)
        lineas = max(lineas, lineas_texto)
    alto = 18 + (lineas - 1) * 15
    return min(alto, 90)


def construir_excel(
    *,
    titulo: str,
    subtitulo: str,
    columnas: List[Columna],
    filas: List[Dict[str, Any]],
    usuario: Any = None,
) -> bytes:
    """Construye un libro Excel en memoria con formato profesional y sutil.

    Args:
        titulo: Titulo del reporte (usado como nombre de hoja).
        subtitulo: Texto secundario.
        columnas: Definicion de columnas.
        filas: Lista de dicts con los datos (claves = ``columna.clave``).
        usuario: Usuario (objeto o dict) que genera el reporte; se muestra en
            los metadatos para dejar trazabilidad del autor.

    Returns:
        bytes con el contenido del archivo .xlsx.
    """
    workbook = Workbook()
    hoja = workbook.active
    # El nombre de la hoja no puede exceder 31 caracteres
    hoja.title = _nombre_hoja(titulo)

    ncols = max(len(columnas), 1)
    # La tabla de datos empieza en la columna B para reservar la A al logo
    offset = OFFSET
    ultima_col_merge = offset + ncols - 1

    # ------------------------------------------------------------------
    # Estilos
    # ------------------------------------------------------------------
    fuente_titulo = Font(name="Calibri", size=17, bold=True, color=COLOR_ROSA_800)
    fuente_sub = Font(name="Calibri", size=10, color=COLOR_TEXTO_SUAVE)
    fuente_meta = Font(name="Calibri", size=9, bold=True, color=COLOR_TEXTO)
    fuente_header = Font(name="Calibri", size=10.5, bold=True, color=COLOR_ROSA_800)
    fuente_celda = Font(name="Calibri", size=10, color=COLOR_TEXTO)

    relleno_metadatos = PatternFill(
        fill_type="solid", start_color=COLOR_ROSA_50, end_color=COLOR_ROSA_50
    )
    relleno_header = PatternFill(
        fill_type="solid", start_color=COLOR_ROSA_100, end_color=COLOR_ROSA_100
    )
    relleno_zebra = PatternFill(fill_type="solid", start_color=COLOR_BLANCO, end_color=COLOR_BLANCO)
    relleno_zebra2 = PatternFill(fill_type="solid", start_color=COLOR_ROSA_50, end_color=COLOR_ROSA_50)

    borde_fino = Side(style="thin", color=COLOR_BORDE)
    borde = Border(left=borde_fino, right=borde_fino, top=borde_fino, bottom=borde_fino)
    borde_marca = Side(style="medium", color=COLOR_ROSA_300)
    borde_inferior_marca = Border(bottom=borde_marca)

    # ------------------------------------------------------------------
    # Logo institucional (Cloudinary) en la esquina superior izquierda
    # ------------------------------------------------------------------
    _agregar_logo(hoja)

    # ------------------------------------------------------------------
    # Encabezado del documento (titulo + subtitulo + metadatos)
    # ------------------------------------------------------------------
    celda_titulo = hoja.cell(row=FILA_TITULO, column=offset, value=titulo)
    celda_titulo.font = fuente_titulo
    celda_titulo.alignment = Alignment(horizontal="left", vertical="center")
    hoja.merge_cells(
        start_row=FILA_TITULO, start_column=offset,
        end_row=FILA_TITULO, end_column=ultima_col_merge,
    )

    celda_sub = hoja.cell(row=FILA_SUBTITULO, column=offset, value=subtitulo)
    celda_sub.font = fuente_sub
    celda_sub.alignment = Alignment(horizontal="left", vertical="center")
    hoja.merge_cells(
        start_row=FILA_SUBTITULO, start_column=offset,
        end_row=FILA_SUBTITULO, end_column=ultima_col_merge,
    )
    hoja.row_dimensions[FILA_TITULO].height = 34
    hoja.row_dimensions[FILA_SUBTITULO].height = 18

    # Metadatos (usuario que genera, fecha y total de registros) en una banda
    # suave de la marca que separa el encabezado del cuerpo.
    datos_usuario = extraer_datos_usuario(usuario)
    ahora = datetime.now(timezone.utc)
    generado = ahora.strftime("%d/%m/%Y %H:%M") + " (UTC)"
    total = len(filas)
    if datos_usuario:
        autor = f"{datos_usuario['nombre']} · {datos_usuario['email']}"
    else:
        autor = "Adoptify"
    meta_texto = (
        f"Generado por: {autor}    ·    Fecha: {generado}    ·    "
        f"Total de registros: {total}"
    )
    celda_meta = hoja.cell(row=FILA_METADATOS, column=offset, value=meta_texto)
    celda_meta.font = fuente_meta
    celda_meta.alignment = Alignment(horizontal="left", vertical="center")
    for j in range(ncols):
        celda_banda = hoja.cell(row=FILA_METADATOS, column=offset + j)
        celda_banda.fill = relleno_metadatos
        celda_banda.border = borde_inferior_marca
    hoja.merge_cells(
        start_row=FILA_METADATOS, start_column=offset,
        end_row=FILA_METADATOS, end_column=ultima_col_merge,
    )
    hoja.row_dimensions[FILA_METADATOS].height = 24

    # ------------------------------------------------------------------
    # Fila de encabezado de columnas
    # ------------------------------------------------------------------
    header_row = FILA_ENCABEZADO
    for idx, col in enumerate(columnas, start=1):
        celda = hoja.cell(row=header_row, column=offset + idx - 1, value=col.titulo)
        celda.font = fuente_header
        celda.fill = relleno_header
        celda.alignment = Alignment(
            horizontal=_ALINEACIONES.get(col.alinear, "left"), vertical="center"
        )
        celda.border = Border(
            left=borde_fino, right=borde_fino, top=borde_fino, bottom=borde_marca
        )
    hoja.row_dimensions[header_row].height = 26

    # ------------------------------------------------------------------
    # Datos
    # ------------------------------------------------------------------
    fila_inicio = header_row + 1
    for i, fila in enumerate(filas):
        fila_excel = fila_inicio + i
        for j, col in enumerate(columnas, start=1):
            valor = _valor_celda(fila, col)
            celda = hoja.cell(row=fila_excel, column=offset + j - 1, value=valor)
            celda.font = fuente_celda
            celda.alignment = Alignment(
                horizontal=_ALINEACIONES.get(col.alinear, "left"),
                vertical="center",
                wrap_text=True,
            )
            celda.border = borde
            if col.tipo in _FORMATOS_EXCEL:
                celda.number_format = _formato_para(col)
            # Alternancia de filas en tonos muy claros de la marca
            celda.fill = relleno_zebra2 if i % 2 == 1 else relleno_zebra

    # ------------------------------------------------------------------
    # Ajustes de la hoja
    # ------------------------------------------------------------------
    ultima_fila = max(fila_inicio + len(filas) - 1, header_row)
    ultima_col = len(columnas)

    # Anchos de columna (clamped a un rango razonable), tabla en columnas B+
    anchos: List[float] = []
    for j, col in enumerate(columnas, start=1):
        ancho = max(min(col.ancho_excel, 40), 8)
        anchos.append(ancho)
        hoja.column_dimensions[get_column_letter(offset + j - 1)].width = ancho

    # Autofiltro sobre el rango de datos
    if filas:
        hoja.auto_filter.ref = (
            f"{get_column_letter(offset)}{header_row}:"
            f"{get_column_letter(offset + ultima_col - 1)}{ultima_fila}"
        )

    # Congelar la fila de encabezado (y el logo) para mantenerlos visibles
    hoja.freeze_panes = f"{get_column_letter(offset)}{fila_inicio}"

    # Altura de filas de datos calculada para no cortar el texto ajustado
    for r in range(fila_inicio, ultima_fila + 1):
        hoja.row_dimensions[r].height = _altura_fila(
            anchos, filas[r - fila_inicio], columnas
        )

    # Vista de impresion (pagina apaisada y ajuste al ancho)
    hoja.page_setup.orientation = "landscape"
    hoja.page_setup.fitToWidth = 1
    hoja.page_setup.fitToHeight = 0
    hoja.sheet_properties.pageSetUpPr.fitToPage = True
    hoja.print_title_rows = f"{header_row}:{header_row}"

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _nombre_hoja(titulo: str) -> str:
    """Normaliza el nombre de la hoja (max 31 chars, sin caracteres invalidos)."""
    limpio = "".join(c if c.isalnum() or c in " -_" else "" for c in titulo).strip()
    return (limpio or "Reporte")[:31]


def _valor_celda(fila: Dict[str, Any], col: Columna):
    """Normaliza el valor a escribir en la celda de Excel."""
    from app.services.reportes.base import _valor_excel

    valor = _valor_excel(fila.get(col.clave), col)
    # Convertir datetime naive a date si solo se usa la parte de fecha
    if isinstance(valor, datetime):
        if col.tipo == TIPO_FECHA and (valor.hour, valor.minute, valor.second) == (0, 0, 0):
            return valor.date()
    return valor
