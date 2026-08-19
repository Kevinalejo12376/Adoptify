"""
Reporte de "Historial de Solicitudes de Adopción" para el rol Usuario.

Genera un PDF profesional (ReportLab) y un libro Excel (openpyxl) con las
adopciones (solicitudes) del usuario autenticado. Reutiliza la paleta de
colores institucional de Adoptify que ya usan los demas reportes del sistema
(``app/services/reportes/pdf.py`` y ``excel.py``) para mantener la identidad
visual en todos los documentos generados.

El PDF incluye:
    - Logo vectorial de Adoptify (huella) + marca.
    - Encabezado repetido en cada pagina.
    - Pie de pagina con numeracion "Pagina X de Y".
    - Titulo, fecha/hora de generacion, nombre del usuario y total de registros.
    - Tabla estilizada con alternancia de filas.

El Excel incluye:
    - Titulo y metadatos (usuario, fecha de generacion).
    - Encabezados en negrilla con colores institucionales.
    - Bordes, formato de fecha, ajuste automatico de columnas y fila congelada.
"""
# pyrefly: ignore [missing-import]
from datetime import datetime, timezone
from io import BytesIO
from typing import Any, Dict, List, Optional

from reportlab.graphics.shapes import Circle, Drawing, Ellipse
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas as _pdf_canvas
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from app.services.reportes.base import obtener_logo_bytes

# ---------------------------------------------------------------------------
# Paleta de colores (identidad Adoptify) - identica a app/services/reportes
# ---------------------------------------------------------------------------
COLOR_PRIMARIO = colors.HexColor("#E11D48")  # Rose-600
COLOR_ACENTO = colors.HexColor("#F59E0B")  # Amber-500
COLOR_ZEBRA = colors.HexColor("#FFF7ED")  # Orange-50
COLOR_TEXTO = colors.HexColor("#1F2937")  # Gray-800
COLOR_TEXTO_SUAVE = colors.HexColor("#6B7280")  # Gray-500
COLOR_BORDE = colors.HexColor("#E5E7EB")  # Gray-200

_ALINEACIONES = {
    "left": TA_LEFT,
    "center": TA_CENTER,
    "right": TA_RIGHT,
}

# Definicion de las columnas del reporte.
# ``clave``: atributo/valor de cada fila; ``titulo``: encabezado visible.
# ``tipo``: "texto" | "fecha"; ``ancho_pdf``: ancho en puntos; ``ancho_excel``:
# ancho minimo en caracteres; ``alinear``: left | center | right.
COLUMNAS: List[Dict[str, Any]] = [
    {"clave": "mascota", "titulo": "Nombre de la mascota", "tipo": "texto",
     "ancho_pdf": 100, "ancho_excel": 22, "alinear": "left"},
    {"clave": "especie", "titulo": "Especie", "tipo": "texto",
     "ancho_pdf": 65, "ancho_excel": 12, "alinear": "center"},
    {"clave": "raza", "titulo": "Raza", "tipo": "texto",
     "ancho_pdf": 85, "ancho_excel": 16, "alinear": "left"},
    {"clave": "refugio", "titulo": "Refugio", "tipo": "texto",
     "ancho_pdf": 130, "ancho_excel": 24, "alinear": "left"},
    {"clave": "fecha", "titulo": "Fecha de adopción", "tipo": "fecha",
     "ancho_pdf": 80, "ancho_excel": 14, "alinear": "center"},
    {"clave": "estado", "titulo": "Estado de la adopción", "tipo": "texto",
     "ancho_pdf": 85, "ancho_excel": 18, "alinear": "center"},
]

_FECHA_ISO = "%d/%m/%Y"


# ===========================================================================
# Datos de la fila
# ===========================================================================
def _fila_valores(solicitud) -> Dict[str, Any]:
    """Extrae los valores de una solicitud en el orden de las columnas.

    Devuelve un dict cuyas claves coinciden con ``clave`` de COLUMNAS.
    """
    mascota = solicitud.mascota
    return {
        "mascota": mascota.nombre if mascota else "—",
        "especie": mascota.tipo.nombre if mascota and mascota.tipo else "—",
        "raza": (mascota.raza or "—") if mascota else "—",
        "refugio": (mascota.refugio.nombre if mascota and mascota.refugio else "—"),
        "fecha": solicitud.creada_en,
        "estado": solicitud.estado.nombre if solicitud.estado else "—",
    }


def _valor_pdf(fila: Dict[str, Any], col: Dict[str, Any]) -> str:
    """Convierte un valor a texto legible para el PDF."""
    valor = fila.get(col["clave"])
    if valor is None:
        return "—"
    if col["tipo"] == "fecha":
        if isinstance(valor, datetime):
            return valor.strftime(_FECHA_ISO)
        return valor.strftime(_FECHA_ISO) if hasattr(valor, "strftime") else str(valor)
    return str(valor)


def _valor_excel(fila: Dict[str, Any], col: Dict[str, Any]):
    """Normaliza el valor para escribirlo en una celda de Excel.

    Las fechas se convierten a ``date`` (sin hora) para que el formato
    ``DD/MM/YYYY`` se aplique de forma limpia.
    """
    valor = fila.get(col["clave"])
    if valor is None:
        return None
    if col["tipo"] == "fecha":
        if isinstance(valor, datetime):
            # Normaliza a fecha (sin hora) para aplicar el formato DD/MM/YYYY
            dt = valor.replace(tzinfo=None) if valor.tzinfo else valor
            if (dt.hour, dt.minute, dt.second) == (0, 0, 0):
                return dt.date()
            return dt
        return valor  # ya es un date
    return valor


# ===========================================================================
# Logo vectorial (huella de Adoptify)
# ===========================================================================
def _dibujar_logo(ancho: float = 40, alto: float = 40) -> Drawing:
    """Dibuja la huella (paw) de Adoptify como logo vectorial del PDF.

    Se dibuja con primitivas de ReportLab para no depender de archivos
    externos ni de conexion a internet al generar el documento.
    """
    color = COLOR_PRIMARIO
    d = Drawing(ancho, alto)
    # Almohadilla principal
    d.add(Ellipse(x=ancho / 2, y=alto * 0.34, rx=ancho * 0.26, ry=alto * 0.19,
                  fillColor=color, strokeColor=None))
    # Cuatro dedos en arco
    d.add(Circle(x=ancho * 0.24, y=alto * 0.70, r=ancho * 0.10, fillColor=color, strokeColor=None))
    d.add(Circle(x=ancho * 0.50, y=alto * 0.80, r=ancho * 0.10, fillColor=color, strokeColor=None))
    d.add(Circle(x=ancho * 0.76, y=alto * 0.70, r=ancho * 0.10, fillColor=color, strokeColor=None))
    d.add(Circle(x=ancho * 0.30, y=alto * 0.56, r=ancho * 0.085, fillColor=color, strokeColor=None))
    d.add(Circle(x=ancho * 0.70, y=alto * 0.56, r=ancho * 0.085, fillColor=color, strokeColor=None))
    return d


# ===========================================================================
# PDF
# ===========================================================================
def _estilos_pdf() -> Dict[str, ParagraphStyle]:
    """Estilos tipograficos del documento PDF."""
    base = getSampleStyleSheet()
    return {
        "titulo": ParagraphStyle(
            "HistorialTitulo", parent=base["Title"], fontName="Helvetica-Bold",
            fontSize=18, leading=22, textColor=COLOR_PRIMARIO, spaceAfter=2,
        ),
        "subtitulo": ParagraphStyle(
            "HistorialSubtitulo", parent=base["Normal"], fontName="Helvetica",
            fontSize=9, leading=12, textColor=COLOR_TEXTO_SUAVE, spaceAfter=0,
        ),
        "meta_clave": ParagraphStyle(
            "HistorialMetaClave", parent=base["Normal"], fontName="Helvetica-Bold",
            fontSize=8.5, leading=12, textColor=COLOR_TEXTO,
        ),
        "meta_valor": ParagraphStyle(
            "HistorialMetaValor", parent=base["Normal"], fontName="Helvetica",
            fontSize=8.5, leading=12, textColor=COLOR_TEXTO_SUAVE,
        ),
        "celda_header": ParagraphStyle(
            "HistorialCeldaHeader", parent=base["Normal"], fontName="Helvetica-Bold",
            fontSize=8.5, leading=11, textColor=colors.white,
        ),
        "vacio": ParagraphStyle(
            "HistorialVacio", parent=base["Normal"], fontName="Helvetica-Oblique",
            fontSize=10, leading=14, textColor=COLOR_TEXTO_SUAVE, alignment=TA_CENTER,
        ),
    }


def _estilo_celda(estilos: Dict[str, ParagraphStyle], alinear: str) -> ParagraphStyle:
    """Devuelve el estilo de celda de datos segun la alineacion."""
    base = estilos["_base"]
    alineacion = _ALINEACIONES.get(alinear, TA_LEFT)
    clave = f"celda_{alinear}"
    if clave not in estilos:
        estilos[clave] = ParagraphStyle(
            clave, parent=base, fontName="Helvetica", fontSize=8, leading=10,
            textColor=COLOR_TEXTO, alignment=alineacion,
        )
    return estilos[clave]


class _CanvasNumerado(_pdf_canvas.Canvas):
    """Canvas que permite dibujar el numero total de paginas al guardar.

    Implementa el patron ``NumberedCanvas`` de ReportLab: guarda el estado de
    cada pagina y al final dibuja el pie con "Pagina X de Y".
    """

    def __init__(self, *args, margen_izq: float = 0, ancho: float = 0, **kwargs):
        super().__init__(*args, **kwargs)
        self._estados_paginas = []
        self._margen_izq = margen_izq
        self._ancho = ancho

    def showPage(self):
        self._estados_paginas.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        total = len(self._estados_paginas)
        for numero, estado in enumerate(self._estados_paginas, start=1):
            self.__dict__.update(estado)
            self._pie_pagina(numero, total)
            super().showPage()
        super().save()

    def _pie_pagina(self, numero: int, total: int):
        """Dibuja el pie de pagina: linea, marca y numeracion."""
        self.saveState()
        margen = self._margen_izq
        ancho = self._ancho
        self.setStrokeColor(COLOR_ACENTO)
        self.setLineWidth(0.6)
        self.line(margen, 12 * mm, ancho - margen, 12 * mm)
        self.setFont("Helvetica", 8)
        self.setFillColor(COLOR_TEXTO_SUAVE)
        self.drawCentredString(ancho / 2, 8 * mm, f"Página {numero} de {total}")
        self.drawString(margen, 8 * mm, "Adoptify · Documento generado automáticamente")
        self.restoreState()


def _canvas_factory(margen_izq: float, ancho: float):
    """Fabrica el canvas numerado inyectando los margenes del documento."""

    def _crear(*args, **kwargs):
        return _CanvasNumerado(*args, margen_izq=margen_izq, ancho=ancho, **kwargs)

    return _crear


def _dibujar_logo_encabezado(canvas, margen: float, alto: float) -> float:
    """Dibuja el logo de Adoptify (desde Cloudinary) en el encabezado.

    Si la imagen no se puede descargar o cargar, usa como respaldo el logo
    vectorial (huella) para que el documento nunca falle.

    Returns:
        Ancho en puntos ocupado por el logo, para posicionar la marca junto a el.
    """
    logo_bytes = obtener_logo_bytes()
    if logo_bytes is not None:
        try:
            from reportlab.lib.utils import ImageReader

            lector = ImageReader(BytesIO(logo_bytes))
            ancho_logo, alto_logo = lector.getSize()
            alto_final = 9 * mm
            ancho_final = alto_final * ancho_logo / alto_logo
            canvas.drawImage(
                BytesIO(logo_bytes),
                margen,
                alto - 17.5 * mm,
                width=ancho_final,
                height=alto_final,
                preserveAspectRatio=True,
                mask="auto",
            )
            return ancho_final
        except Exception:
            # Fallo de red/carga: se usa el logo vectorial de respaldo.
            pass

    logo = _dibujar_logo()
    logo.drawOn(canvas, margen, alto - 18 * mm)
    return 9 * mm


def _dibujar_encabezado(canvas, doc):
    """Dibuja el encabezado de cada pagina: barra, logo, marca y linea de acento."""
    canvas.saveState()
    ancho = doc.pagesize[0]
    alto = doc.pagesize[1]
    margen = doc.leftMargin

    # Barra superior institucional
    canvas.setFillColor(COLOR_PRIMARIO)
    canvas.rect(0, alto - 2 * mm, ancho, 2 * mm, stroke=0, fill=1)

    # Logo (Cloudinary) + marca Adoptify
    ancho_logo = _dibujar_logo_encabezado(canvas, margen, alto)
    x_marca = margen + ancho_logo + 4 * mm
    canvas.setFont("Helvetica-Bold", 12)
    canvas.setFillColor(COLOR_PRIMARIO)
    canvas.drawString(x_marca, alto - 15 * mm, "ADOPTIFY")
    canvas.setFont("Helvetica", 7.5)
    canvas.setFillColor(COLOR_TEXTO_SUAVE)
    canvas.drawString(x_marca, alto - 18 * mm, "Plataforma de adopción de mascotas")

    # Seccion a la derecha
    canvas.setFont("Helvetica-Bold", 9)
    canvas.setFillColor(COLOR_TEXTO)
    canvas.drawRightString(ancho - margen, alto - 15 * mm, "Reporte del Usuario")

    # Linea de acento
    canvas.setStrokeColor(COLOR_ACENTO)
    canvas.setLineWidth(0.8)
    canvas.line(margen, alto - 21 * mm, ancho - margen, alto - 21 * mm)
    canvas.restoreState()


def _tabla_meta(usuario, total: int, generado: str) -> Table:
    """Tabla compacta con los metadatos: usuario, correo, total y fecha."""
    estilos = _estilos_pdf()
    nombre = f"{usuario.nombre} {usuario.apellido or ''}".strip() or "—"
    datos = [
        [
            Paragraph("Usuario", estilos["meta_clave"]),
            Paragraph(nombre, estilos["meta_valor"]),
            Paragraph("Total de registros", estilos["meta_clave"]),
            Paragraph(str(total), estilos["meta_valor"]),
        ],
        [
            Paragraph("Correo", estilos["meta_clave"]),
            Paragraph(usuario.email or "—", estilos["meta_valor"]),
            Paragraph("Generado", estilos["meta_clave"]),
            Paragraph(generado, estilos["meta_valor"]),
        ],
    ]
    tabla = Table(datos, colWidths=[30 * mm, 65 * mm, 35 * mm, 50 * mm])
    tabla.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LINEBELOW", (0, 0), (-1, -2), 0.4, COLOR_BORDE),
        ("BOX", (0, 0), (-1, -1), 0.8, COLOR_PRIMARIO),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#FFF1F2")),  # Rose-50
    ]))
    return tabla


def _estilo_tabla(n_filas: int) -> TableStyle:
    """Estilo visual profesional de la tabla de datos."""
    commands = [
        ("BACKGROUND", (0, 0), (-1, 0), COLOR_PRIMARIO),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8.5),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("TEXTCOLOR", (0, 1), (-1, -1), COLOR_TEXTO),
        ("GRID", (0, 0), (-1, -1), 0.4, COLOR_BORDE),
        ("BOX", (0, 0), (-1, -1), 0.8, COLOR_PRIMARIO),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("LINEBELOW", (0, 0), (-1, 0), 1.2, COLOR_ACENTO),
    ]
    # Alternancia de filas (zebra)
    for r in range(1, n_filas + 1):
        if r % 2 == 0:
            commands.append(("BACKGROUND", (0, r), (-1, r), COLOR_ZEBRA))
    return TableStyle(commands)


def generar_pdf_historial(usuario, solicitudes: List[Any]) -> bytes:
    """Construye el PDF "Historial de Solicitudes de Adopción" en memoria.

    Args:
        usuario: Objeto Usuario autenticado (para nombre y correo).
        solicitudes: Lista de SolicitudAdopcion del usuario.

    Returns:
        bytes con el contenido del PDF.
    """
    estilos = _estilos_pdf()
    estilos["_base"] = getSampleStyleSheet()["Normal"]
    buffer = BytesIO()

    # Orientacion horizontal para mejor legibilidad de las 6 columnas
    pagina = landscape(A4)
    margen = 12 * mm
    doc = SimpleDocTemplate(
        buffer,
        pagesize=pagina,
        leftMargin=margen,
        rightMargin=margen,
        topMargin=22 * mm,
        bottomMargin=16 * mm,
        title="Historial de Solicitudes de Adopción",
        author="Adoptify",
        subject="Historial de solicitudes de adopción del usuario",
        onPage=_dibujar_encabezado,
        canvasmaker=_canvas_factory(margen, pagina[0]),
    )

    ahora = datetime.now(timezone.utc)
    generado = ahora.strftime("%d/%m/%Y %H:%M") + " (UTC)"
    total = len(solicitudes)

    elementos = [
        Paragraph("Historial de Solicitudes de Adopción", estilos["titulo"]),
        Paragraph(
            f"Generado el {generado} · Adoptify", estilos["subtitulo"]
        ),
        Spacer(1, 5 * mm),
        _tabla_meta(usuario, total, generado),
        Spacer(1, 6 * mm),
    ]

    if not solicitudes:
        elementos.append(
            Paragraph(
                "No tienes solicitudes de adopción registradas.", estilos["vacio"]
            )
        )
    else:
        filas = [_fila_valores(s) for s in solicitudes]
        data = [
            [Paragraph(c["titulo"], estilos["celda_header"]) for c in COLUMNAS]
        ]
        for fila in filas:
            data.append(
                [
                    Paragraph(
                        _valor_pdf(fila, col),
                        _estilo_celda(estilos, col["alinear"]),
                    )
                    for col in COLUMNAS
                ]
            )

        ancho_util = doc.width
        anchos = [c["ancho_pdf"] for c in COLUMNAS]
        if sum(anchos) > ancho_util:
            factor = ancho_util / sum(anchos)
            anchos = [max(a * factor, 40) for a in anchos]

        tabla = Table(data, colWidths=anchos, repeatRows=1)
        tabla.setStyle(_estilo_tabla(len(filas)))
        elementos.append(tabla)

    doc.build(elementos)
    return buffer.getvalue()


# ===========================================================================
# Excel
# ===========================================================================
def _autoajustar_columnas(hoja, filas: List[Any], header_row: int):
    """Ajusta automaticamente el ancho de cada columna segun su contenido."""
    from openpyxl.utils import get_column_letter

    for j, col in enumerate(COLUMNAS, start=1):
        ancho = len(col["titulo"]) + 2
        for solicitud in filas:
            fila = _fila_valores(solicitud)
            valor = fila.get(col["clave"])
            if isinstance(valor, str):
                ancho = max(ancho, len(valor) + 2)
        # Las fechas necesitan espacio para DD/MM/YYYY
        if col["tipo"] == "fecha":
            ancho = max(ancho, 12)
        hoja.column_dimensions[get_column_letter(j)].width = min(max(ancho, 10), 45)


def generar_excel_historial(usuario, solicitudes: List[Any]) -> bytes:
    """Construye el libro Excel "Historial de Solicitudes de Adopción" en memoria.

    Args:
        usuario: Objeto Usuario autenticado (para nombre y correo).
        solicitudes: Lista de SolicitudAdopcion del usuario.

    Returns:
        bytes con el contenido del archivo .xlsx.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    color_primario = "E11D48"  # Rose-600
    color_zebra = "FFF7ED"  # Orange-50
    color_texto = "1F2937"
    color_texto_suave = "6B7280"
    color_borde = "E5E7EB"

    workbook = Workbook()
    hoja = workbook.active
    hoja.title = "Historial de Solicitudes"

    fuente_titulo = Font(name="Calibri", size=14, bold=True, color=color_primario)
    fuente_sub = Font(name="Calibri", size=9, italic=True, color=color_texto_suave)
    fuente_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    relleno_header = PatternFill(fill_type="solid", start_color=color_primario, end_color=color_primario)
    relleno_zebra = PatternFill(fill_type="solid", start_color=color_zebra, end_color=color_zebra)
    fuente_celda = Font(name="Calibri", size=10, color=color_texto)
    borde_fino = Side(style="thin", color=color_borde)
    borde = Border(left=borde_fino, right=borde_fino, top=borde_fino, bottom=borde_fino)

    ncols = len(COLUMNAS)
    nombre_usuario = f"{usuario.nombre} {usuario.apellido or ''}".strip() or "—"
    ahora = datetime.now(timezone.utc)
    generado = ahora.strftime("%d/%m/%Y %H:%M") + " (UTC)"

    # ------------------------------------------------------------------
    # Logo institucional (Cloudinary) en la esquina superior izquierda
    # ------------------------------------------------------------------
    from app.services.reportes.excel import _agregar_logo

    _agregar_logo(hoja)

    # ------------------------------------------------------------------
    # Encabezado del documento (titulo + subtitulo)
    # ------------------------------------------------------------------
    hoja.cell(
        row=1, column=2, value="Historial de Solicitudes de Adopción"
    ).font = fuente_titulo
    hoja.cell(
        row=2, column=2,
        value=f"Adoptify · Usuario: {nombre_usuario} · Generado: {generado}",
    ).font = fuente_sub
    hoja.merge_cells(start_row=1, start_column=2, end_row=1, end_column=ncols + 1)
    hoja.merge_cells(start_row=2, start_column=2, end_row=2, end_column=ncols + 1)
    hoja.row_dimensions[1].height = 26
    hoja.row_dimensions[2].height = 14
    hoja.column_dimensions["A"].width = 16

    # ------------------------------------------------------------------
    # Fila de encabezado de columnas (fila 4)
    # ------------------------------------------------------------------
    header_row = 4
    for j, col in enumerate(COLUMNAS, start=1):
        celda = hoja.cell(row=header_row, column=j, value=col["titulo"])
        celda.font = fuente_header
        celda.fill = relleno_header
        celda.alignment = Alignment(
            horizontal=_ALINEACIONES_EXCEL.get(col["alinear"], "left"),
            vertical="center",
        )
        celda.border = borde
    hoja.row_dimensions[header_row].height = 20

    # ------------------------------------------------------------------
    # Datos
    # ------------------------------------------------------------------
    fila_inicio = header_row + 1
    for i, solicitud in enumerate(solicitudes):
        fila_excel = fila_inicio + i
        fila = _fila_valores(solicitud)
        for j, col in enumerate(COLUMNAS, start=1):
            celda = hoja.cell(row=fila_excel, column=j, value=_valor_excel(fila, col))
            celda.font = fuente_celda
            celda.alignment = Alignment(
                horizontal=_ALINEACIONES_EXCEL.get(col["alinear"], "left"),
                vertical="center",
                wrap_text=True,
            )
            celda.border = borde
            if col["tipo"] == "fecha":
                celda.number_format = "DD/MM/YYYY"
            if i % 2 == 1:
                celda.fill = relleno_zebra
        hoja.row_dimensions[fila_excel].height = 18

    # ------------------------------------------------------------------
    # Ajustes de la hoja
    # ------------------------------------------------------------------
    _autoajustar_columnas(hoja, solicitudes, header_row)

    ultima_fila = max(fila_inicio + len(solicitudes) - 1, header_row)
    ultima_col = ncols
    if solicitudes:
        hoja.auto_filter.ref = f"A{header_row}:{get_column_letter(ultima_col)}{ultima_fila}"
    # Congela el encabezado (y titulo) para mantenerlo visible al desplazarse
    hoja.freeze_panes = f"A{fila_inicio}"

    # Vista de impresion
    hoja.page_setup.orientation = "landscape"
    hoja.page_setup.fitToWidth = 1
    hoja.page_setup.fitToHeight = 0
    hoja.sheet_properties.pageSetUpPr.fitToPage = True
    hoja.print_title_rows = f"{header_row}:{header_row}"

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


# Alineaciones de Excel (se reutiliza el mapeo de la identidad visual)
_ALINEACIONES_EXCEL = {
    "left": "left",
    "center": "center",
    "right": "right",
}


# Exposicion util para pruebas/uso desde otros modulos
__all__ = [
    "COLUMNAS",
    "generar_pdf_historial",
    "generar_excel_historial",
]
